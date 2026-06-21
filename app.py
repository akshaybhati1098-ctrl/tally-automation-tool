import os
from dotenv import load_dotenv
from fastapi import Request
load_dotenv()
import io
import uuid
import logging
import json
import bcrypt
import secrets
import tempfile
import time       # Added for tracking execution time metric values
import traceback  # Added for capturing crash logs safely
import asyncio
from concurrent.futures import ThreadPoolExecutor

MATCH_SESSIONS = {}
JOBS = {}
RESULTS = {}
# Job statuses: PENDING, PROCESSING, COMPLETED, FAILED
JOB_STATUS = {}  # {job_id: {"status": "...", "progress": 0-100, "message": "..."}}

# Thread pool for blocking operations
thread_pool = ThreadPoolExecutor(max_workers=4)

from datetime import datetime, timedelta
os.makedirs("data", exist_ok=True)
from io import BytesIO
from typing import Optional

from fastapi import FastAPI, Request, UploadFile, Form, HTTPException, Depends, File, WebSocket, WebSocketDisconnect
from fastapi.responses import Response, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware
from fastapi.middleware.cors import CORSMiddleware
from core.excel_service import (
    excel_to_xml,
    prepare_excel_party_matching,
    apply_corrections_and_build_final_df,
    export_dataframe_to_excel_bytes,
)
from core.tally_service import (
    build_ledger_xml,
    build_company_status_xml,
    parse_ledgers,
    parse_company_status,
)
from core.match_service import apply_match_results_to_dataframe

import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

# PostgreSQL
import psycopg2
from psycopg2.extras import RealDictCursor

# Email verification imports
from itsdangerous import BadSignature, SignatureExpired
from core.email import generate_token, decode_token, send_verification_email, send_otp_email, send_welcome_email

# Existing services – these now use persistent /data/mapping.json
from core.excel_service import excel_to_xml
from core.mapping import (
    load_companies,
    add_company,
    delete_company,
    get_company_mapping as get_company_mapping_data,
    save_company_mapping
)
from core.process_service import image_to_excel

# New Admin Enterprise Layer Dependencies
from core.admin_telemetry import log_admin_event, ensure_admin_schema
from routes.admin_routes import admin_router as enterprise_admin_system_router

# --- NEW BUSINESS EVENT TELEMETRY DEPENDENCIES ---
from core.business_telemetry import (
    log_match_event,
    log_conversion_event,
    log_ocr_event,
    log_business_error
)
from routes.admin_business_routes import business_router

# =========================================================
# APP INITIALIZATION
# =========================================================
app = FastAPI(title="Tally Automation Tool")

# =========================================================
# MIDDLEWARE STACK ORDER (Fixed for Session Context)
# =========================================================

# 1. CORS Settings
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://tallytool.online",
        "https://www.tallytool.online"
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 2. Enterprise Telemetry Instrumentation Middleware
@app.middleware("http")
async def enterprise_request_instrumentation_middleware(request: Request, call_next):
    # Skip tracking static assets to protect database performance and storage space
    if request.url.path.startswith("/static") or request.url.path.startswith("/favicon.ico") or request.url.path.startswith("/admin"):
        return await call_next(request)

    start_time = time.perf_counter()
    
    try:
        user_id = request.session.get("user_id")
        username = request.session.get("username") or "anonymous_guest"
    except AssertionError:
        user_id = None
        username = "anonymous_guest"
        
    endpoint_path = request.url.path
    
    try:
        response = await call_next(request)
        duration_ms = int((time.perf_counter() - start_time) * 1000)
        
        # Track core analytical traffic pipelines and all API transactions dynamically
        if endpoint_path.startswith("/api/") or endpoint_path in ["/login", "/signup"]:
            log_admin_event(
                user_id=user_id,
                username=username,
                event_type="api_request",
                endpoint=endpoint_path,
                status_str="success" if response.status_code < 400 else "failed",
                execution_time_ms=duration_ms,
                details={"status_code": response.status_code, "method": request.method}
            )
        return response

    except Exception as unhandled_sys_exc:
        duration_ms = int((time.perf_counter() - start_time) * 1000)
        error_msg = str(unhandled_sys_exc)
        stack_trace = traceback.format_exc()
        
        # Record structural application errors completely down to the database row
        log_admin_event(
            user_id=user_id,
            username=username,
            event_type="system_crash",
            endpoint=endpoint_path,
            status_str="error",
            error_message=error_msg,
            execution_time_ms=duration_ms,
            details={"stack_trace": stack_trace, "method": request.method}
        )
        raise unhandled_sys_exc
    
@app.middleware("http")
async def admin_security_no_cache_middleware(request: Request, call_next):
    """Forces the browser to delete the admin page from memory the second you log out."""
    response = await call_next(request)
    
    if request.url.path.startswith("/admin"):
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
        
    return response

# 3. Session Timeout Tracker Middleware
SESSION_TIMEOUT_MINUTES = 60

@app.middleware("http")
async def session_timeout_middleware(request: Request, call_next):
    if "session" in request.scope:
        try:
            if request.session.get("username"):
                last_active = request.session.get("last_active")
                if last_active:
                    last_dt = datetime.fromisoformat(last_active)
                    if datetime.now() - last_dt > timedelta(minutes=SESSION_TIMEOUT_MINUTES):
                        request.session.clear()
                request.session["last_active"] = datetime.now().isoformat()
        except Exception:
            pass
    return await call_next(request)

# 4. Session Engine Core Initialization Middleware
SECRET_KEY = os.environ.get("SECRET_KEY")
if not SECRET_KEY:
    raise RuntimeError("SECRET_KEY environment variable not set")
app.add_middleware(
    SessionMiddleware,
    secret_key=SECRET_KEY
)

# =========================================================
# STATIC & TEMPLATES
# =========================================================
app.mount("/static", StaticFiles(directory="web/static"), name="static")
templates = Jinja2Templates(directory="web/templates")

# =========================================================
# USER DB (PostgreSQL) – persistent across rebuilds
# =========================================================
DATABASE_URL = os.environ.get("DATABASE_URL")
if not DATABASE_URL:
    raise RuntimeError("DATABASE_URL environment variable not set")

def get_db_connection(retries=3, backoff=1):
    """Connect to database with retry logic for transient failures."""
    for attempt in range(retries):
        try:
            return psycopg2.connect(
                DATABASE_URL,
                sslmode="require",
                connect_timeout=10
            )
        except psycopg2.OperationalError as e:
            if attempt < retries - 1:
                wait_time = backoff * (2 ** attempt)
                print(f"⚠️ DB connection attempt {attempt + 1} failed: {str(e)[:100]}")
                print(f"   Retrying in {wait_time}s...")
                time.sleep(wait_time)
            else:
                print(f"❌ DB connection failed after {retries} attempts")
                raise

def init_user_db():
    """Create users and pending_users tables if they don't exist."""
    conn = get_db_connection()
    cur = conn.cursor()

    # Initialize Telemetry Schema Immediately to Avoid Database Race Conditions
    try:
        ensure_admin_schema()
        print("✅ Admin telemetry logging schema verified/created.")
    except Exception as telemetry_init_err:
        print(f"⚠️ Warning: Pre-initializing admin schema failed: {telemetry_init_err}")

    # Users table (final) with email and verification fields
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY,
            username TEXT UNIQUE NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            is_verified INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # Pending registrations (temporary) – stores OTP before final signup
    cur.execute("""
        CREATE TABLE IF NOT EXISTS pending_users (
            email TEXT PRIMARY KEY,
            username TEXT NOT NULL,
            otp_code TEXT NOT NULL,
            otp_expiry TIMESTAMP NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # Add reset token columns to users table if they don't exist
    cur.execute("SELECT column_name FROM information_schema.columns WHERE table_name='users'")
    columns = [col[0] for col in cur.fetchall()]

    if 'reset_token' not in columns:
        cur.execute("ALTER TABLE users ADD COLUMN reset_token TEXT")
        print("Added reset_token column to users table")

    if 'reset_expiry' not in columns:
        cur.execute("ALTER TABLE users ADD COLUMN reset_expiry TIMESTAMP")
        print("Added reset_expiry column to users table")

    if 'is_admin' not in columns:
        cur.execute("ALTER TABLE users ADD COLUMN is_admin BOOLEAN DEFAULT FALSE")
        print("Added is_admin column to users table")

    # =========================================================
    # AUTOMATIC TELEMETRY COLUMN MIGRATE & FIX
    # =========================================================
    try:
        # 1. Scan for all existing tables in the active database
        cur.execute("""
            SELECT table_name 
            FROM information_schema.tables 
            WHERE table_schema = 'public'
        """)
        existing_tables = [row[0] for row in cur.fetchall()]
        print(f"🔍 Connected Database Tables Found: {existing_tables}")
        
        # 2. Force add 'execution_time_ms' to any table that looks like a telemetry/logging table
        candidates = ['admin_events', 'telemetry', 'logs', 'api_requests', 'admin_telemetry']
        for table in existing_tables:
            if table in candidates or 'log' in table or 'event' in table or 'telemetry' in table:
                try:
                    cur.execute(f"ALTER TABLE {table} ADD COLUMN IF NOT EXISTS execution_time_ms INTEGER DEFAULT 0;")
                    print(f"✅ Verified/Added 'execution_time_ms' to table: {table}")
                except Exception as table_err:
                    print(f"⚠️ Skip table update for {table}: {table_err}")
    except Exception as migration_err:
        print(f"❌ Automatic telemetry migration error: {migration_err}")
    # =========================================================

    conn.commit()
    cur.close()
    conn.close()
    print("✅ User database tables initialized in PostgreSQL")

init_user_db()

# =========================================================
# USER HELPERS (existing + new)
# =========================================================
def get_user(username: str):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("SELECT * FROM users WHERE username = %s", (username,))
    user = cur.fetchone()
    cur.close()
    conn.close()
    return user

def get_user_by_email(email: str):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("SELECT * FROM users WHERE email = %s", (email,))
    user = cur.fetchone()
    cur.close()
    conn.close()
    return user

# ================= FIXED create_user =================
def create_user(username: str, email: str, password: str):
    hashed = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute(
        "INSERT INTO users (username, email, password_hash) VALUES (%s, %s, %s) RETURNING id",
        (username, email, hashed)
    )

    user_id = cur.fetchone()[0]

    conn.commit()
    cur.close()
    conn.close()

    print(f"✅ User '{username}' created")
    return user_id

# Legacy function (username only) – kept for backward compatibility if needed
def create_user_legacy(username: str, password: str):
    hashed = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO users (username, email, password_hash) VALUES (%s, %s, %s)",
        (username, f"{username}@temp.local", hashed)  # placeholder email
    )
    conn.commit()
    cur.close()
    conn.close()
    print(f"✅ User '{username}' created in PostgreSQL (legacy, with placeholder email)")

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode(), hashed.encode())

# =========================================================
# PASSWORD RESET HELPERS (NEW)
# =========================================================
def set_user_reset_token(email: str, token: str, expiry: datetime):
    """Store reset token and expiry for user."""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "UPDATE users SET reset_token = %s, reset_expiry = %s WHERE email = %s",
        (token, expiry, email)
    )
    conn.commit()
    cur.close()
    conn.close()
    print(f"✅ Reset token set for {email}")

def get_user_by_reset_token(token: str):
    """Get user by valid reset token."""
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute(
        "SELECT * FROM users WHERE reset_token = %s AND reset_expiry > NOW()",
        (token,)
    )
    user = cur.fetchone()
    cur.close()
    conn.close()
    return user

def clear_reset_token(email: str):
    """Clear reset token after use."""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "UPDATE users SET reset_token = NULL, reset_expiry = NULL WHERE email = %s",
        (email,)
    )
    conn.commit()
    cur.close()
    conn.close()

def update_user_password(email: str, new_password: str):
    """Update user's password (hashed)."""
    hashed = bcrypt.hashpw(new_password.encode(), bcrypt.gensalt()).decode()
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "UPDATE users SET password_hash = %s WHERE email = %s",
        (hashed, email)
    )
    conn.commit()
    cur.close()
    conn.close()
    print(f"✅ Password updated for {email}")

# =========================================================
# PENDING USER HELPERS (OTP)
# =========================================================
def save_pending_user(email: str, username: str, otp: str, expiry: datetime):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO pending_users (email, username, otp_code, otp_expiry)
        VALUES (%s, %s, %s, %s)
        ON CONFLICT (email) DO UPDATE SET
            username = EXCLUDED.username,
            otp_code = EXCLUDED.otp_code,
            otp_expiry = EXCLUDED.otp_expiry
    """, (email, username, otp, expiry))
    conn.commit()
    cur.close()
    conn.close()

def get_pending_user(email: str):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("SELECT * FROM pending_users WHERE email = %s", (email,))
    user = cur.fetchone()
    cur.close()
    conn.close()
    return user

def delete_pending_user(email: str):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM pending_users WHERE email = %s", (email,))
    conn.commit()
    cur.close()
    conn.close()

# =========================================================
# AUTH HELPERS
# =========================================================
def get_current_user(request: Request) -> Optional[str]:
    username = request.session.get("username")
    if not username:
        return None
        
    # 🔥 LIVE SECURITY CHECK: Verify they aren't suspended right now
    user = get_user(username)
    if not user or user.get("is_active") is False:
        request.session.clear() # Instantly destroy their session
        return None
        
    return username

def require_login(request: Request):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    return user

def require_admin(request: Request):
    """Route guard enforcing that the current session belongs to a platform administrator."""
    username = get_current_user(request)
    if not username:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    user = get_user(username)
    if not user or not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Administrative access clearance denied")
    return username

def get_session_user_id(request: Request) -> str:
    """Return logged-in user id as string for connector queues."""
    user_id = request.session.get("user_id")
    if user_id is None:
        raise HTTPException(status_code=401, detail="Not authenticated")
    return str(user_id)

# =========================================================
# UI ROUTES
# =========================================================
@app.get("/")
async def serve_ui(request: Request):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login")
    return templates.TemplateResponse(
        "index.html",
        {"request": request, "username": user}
    )
# Device-scoped connector status (Tally on this machine)
CONNECTOR_STATUS = {}


def get_device_id_from_request(request: Request) -> str:
    device_id = (
        request.headers.get("X-Device-ID")
        or request.headers.get("X-Device-Id")
        or request.headers.get("x-device-id")
    )
    if not device_id or not str(device_id).strip():
        raise HTTPException(
            status_code=400,
            detail="Missing X-Device-ID header.",
        )
    return str(device_id).strip()


def _connector_status_payload(device_id: str) -> dict:
    data = CONNECTOR_STATUS.get(device_id)
    if isinstance(data, dict):
        return {
            "status": data.get("status", "not_running"),
            "company": data.get("company"),
        }
    return {
        "status": data or "not_running",
        "company": None,
    }

def _parse_last_seen(value) -> Optional[datetime]:
    """Parse heartbeat timestamp; missing/invalid values sort as oldest."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        return None


def _connector_last_seen_sort_key(item: tuple) -> datetime:
    """Never use raw string max() — missing last_seen must not win over real heartbeats."""
    _device_id, payload = item
    if not isinstance(payload, dict):
        return datetime.min
    parsed = _parse_last_seen(payload.get("last_seen"))
    return parsed if parsed is not None else datetime.min


def _get_latest_connector_entry() -> tuple[Optional[str], Optional[dict]]:
    if not CONNECTOR_STATUS:
        return None, None
    device_id, payload = max(CONNECTOR_STATUS.items(), key=_connector_last_seen_sort_key)
    return device_id, payload if isinstance(payload, dict) else None


@app.post("/api/connector/heartbeat/{device_id}")
def connector_heartbeat(device_id: str, data: dict):
    """Connector posts Tally running state for this PC/device."""
    status = data.get("status")
    company = data.get("company")
    
    # 🔥 FIX: Intercept the user identity from the connector
    username = data.get("username")
    user_id = data.get("user_id")
    
    
    # Save the identity into the active memory dictionary
    CONNECTOR_STATUS[device_id] = {
        "status": status,
        "company": company,
        "username": username,
        "user_id": user_id,
        "last_seen": datetime.now().isoformat(),
    }
    return {"success": True}

import time

@app.get("/api/connector/status")
def connector_status(request: Request):
    start = time.time()

    device_id = get_device_id_from_request(request)
    status_data = _connector_status_payload(device_id)

    print(f"connector_status took {time.time()-start:.3f}s")

    return JSONResponse(status_data)

@app.get("/login")
async def login_page(request: Request):
    return templates.TemplateResponse("pages/login.html", {"request": request})

@app.post("/login")
async def login_post(
    request: Request,
    username: str = Form(...),
    password: str = Form(...)
):
    username = username.strip()
    user = get_user(username)
    
    # 1. 🔥 SECURITY CHECK: Intercept Suspended Accounts Immediately
    if user and user.get("is_active") is False:
        # Use your existing flash alert system from login.html
        return templates.TemplateResponse("pages/login.html", {
            "request": request,
            "flashes": [{"category": "error", "message": "Your account has been suspended. Please contact support."}]
        })

    # 2. Verify Password and Proceed
    if user and verify_password(password, user["password_hash"]):
        request.session["username"] = username
        request.session["user_id"] = user["id"]   
        
        # --- BUSINESS EVENT LOGGING: SUCCESS ---
        log_admin_event(
            user_id=user["id"],
            username=username,
            event_type="login_success",
            status_str="success",
            details={
                "username": username,
                "ip_address": request.client.host,
                "user_agent": request.headers.get("user-agent", "unknown")
            }
        )
        return RedirectResponse("/", status_code=302)
        
    # --- BUSINESS EVENT LOGGING: FAILURE ---
    log_admin_event(
        username=username,
        event_type="login_failed",
        status_str="failed",
        details={
            "username": username,
            "ip_address": request.client.host,
            "user_agent": request.headers.get("user-agent", "unknown")
        }
    )
    log_business_error(
        user_id=None,
        username=username,
        event_type="login",
        error_type="auth_failure",
        error_message="Invalid account handle credentials combination input matching query signature values."
    )
    return RedirectResponse("/login?error=1", status_code=302)
        
    # --- BUSINESS EVENT LOGGING: FAILURE ---
    log_admin_event(
        username=username,
        event_type="login_failed",
        status_str="failed",
        details={
            "username": username,
            "ip_address": request.client.host,
            "user_agent": request.headers.get("user-agent", "unknown")
        }
    )
    log_business_error(
        user_id=None,
        username=username,
        event_type="login",
        error_type="auth_failure",
        error_message="Invalid account handle credentials combination input matching query signature values."
    )
    return RedirectResponse("/login?error=1", status_code=302)

@app.get("/signup")
async def signup_page(request: Request):
    return templates.TemplateResponse("pages/signup.html", {"request": request})

# Legacy signup endpoint (username/password only) – kept for backward compatibility
@app.post("/signup")
async def signup_post(
    request: Request,
    username: str = Form(...),
    password: str = Form(...)
):
    username = username.strip()
    if not username or not password:
        return RedirectResponse("/signup?error=1", status_code=302)
    if get_user(username):
        return RedirectResponse("/signup?exists=1", status_code=302)
    create_user_legacy(username, password)
    return RedirectResponse("/login?created=1", status_code=302)

@app.get("/logout")
async def logout(request: Request):
    user = get_current_user(request)
    if user:
        # --- BUSINESS EVENT LOGGING ---
        log_admin_event(
            username=user,
            event_type="logout",
            status_str="success"
        )
        
    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
    request.session.clear()
    if is_ajax:
        return JSONResponse({"success": True, "redirect": "/login"})
    return RedirectResponse("/login")

@app.get("/api/me")
async def api_me(request: Request):
    user = get_current_user(request)
    user_id = request.session.get("user_id")
    return {
        "authenticated": bool(user),
        "username": user,
        "user_id": user_id,
    }

# =========================================================
# OTP ENDPOINTS (for email verification signup)
# =========================================================
@app.post("/api/send-otp")
async def send_otp(email: str = Form(...), username: str = Form(...)):
    """Generate and email a 6-digit OTP for signup."""
    email = email.strip().lower()
    username = username.strip()

    # Basic validation
    if not email or not username:
        return JSONResponse({"status": "error", "message": "Email and username required."}, status_code=400)

    # Check if email already registered
    if get_user_by_email(email):
        return JSONResponse({"status": "error", "message": "Email already registered."}, status_code=400)

    # Check if username already exists
    if get_user(username):
        return JSONResponse({"status": "error", "message": "Username already taken."}, status_code=400)

    # Overwrite any existing pending record for this email
    if get_pending_user(email):
        delete_pending_user(email)

    # Generate 6-digit OTP
    otp = f"{secrets.randbelow(1000000):06d}"
    expiry = datetime.now() + timedelta(minutes=10)  # OTP valid for 10 minutes

    # Store pending record
    save_pending_user(email, username, otp, expiry)

    # Send OTP email
    try:
        send_otp_email(email, otp)
    except Exception as e:
        print(f"OTP email failed: {e}")
        delete_pending_user(email)
        return JSONResponse({"status": "error", "message": "Failed to send OTP email. Please check your email address or try again later."}, status_code=500)

    return JSONResponse({"status": "ok"})

@app.post("/api/verify-otp-signup")
async def verify_otp_signup(
    email: str = Form(...),
    username: str = Form(...),
    password: str = Form(...),
    otp: str = Form(...)
):
    email = email.strip().lower()
    username = username.strip()

    pending = get_pending_user(email)
    if not pending:
        return JSONResponse({"status": "error", "message": "No pending registration found"}, status_code=400)

    if pending["username"] != username:
        return JSONResponse({"status": "error", "message": "Username mismatch"}, status_code=400)

    if pending["otp_code"] != otp:
        return JSONResponse({"status": "error", "message": "Invalid OTP"}, status_code=400)

    expiry = pending["otp_expiry"]
    if isinstance(expiry, str):
        expiry = datetime.fromisoformat(expiry)

    if datetime.now() > expiry:
        delete_pending_user(email)
        return JSONResponse({"status": "error", "message": "OTP expired"}, status_code=400)

    # Final validation
    if get_user_by_email(email) or get_user(username):
        delete_pending_user(email)
        return JSONResponse({"status": "error", "message": "User already exists"}, status_code=400)

    # CREATE USER
    user_id = create_user(username, email, password)

    # DEFAULT COMPANY
    from core.mapping import save_company_mapping_postgres, get_default_mapping
    save_company_mapping_postgres("Default", get_default_mapping(), user_id)

    # SEND EMAIL (optional)
    try:
        send_welcome_email(email, username)
    except Exception as e:
        print("Email error:", e)

    # DELETE PENDING USER
    delete_pending_user(email)

    return JSONResponse({"status": "ok"})

# =========================================================
# FORGOT USERNAME/PASSWORD ENDPOINTS (NEW)
# =========================================================
@app.post("/api/forgot-username")
async def forgot_username(email: str = Form(...)):
    """Send username to user's email."""
    email = email.strip().lower()
    user = get_user_by_email(email)
    if not user:
        # Return success even if email not found (security)
        return JSONResponse({"status": "ok"})
    
    try:
        # Import the email function
        from core.email import send_username_reminder_email
        send_username_reminder_email(email, user["username"])
    except Exception as e:
        print(f"Failed to send username reminder: {e}")
        return JSONResponse({"status": "error", "message": "Failed to send email."}, status_code=500)
    
    return JSONResponse({"status": "ok"})

@app.post("/api/forgot-password")
async def forgot_password(email: str = Form(...)):
    """Send password reset link to user's email."""
    email = email.strip().lower()
    user = get_user_by_email(email)
    if not user:
        return JSONResponse({"status": "ok"})  # Security: don't reveal existence
    
    # Generate reset token (valid for 1 hour)
    from core.email import generate_token
    reset_token = generate_token(email)
    expiry = datetime.now() + timedelta(hours=1)
    
    set_user_reset_token(email, reset_token, expiry)
    
    BASE_URL = os.environ.get("BASE_URL", "https://tallytool.online")
    reset_link = f"{BASE_URL}/reset-password?token={reset_token}"
    
    try:
        from core.email import send_password_reset_email
        send_password_reset_email(email, reset_link)
    except Exception as e:
        print(f"Failed to send password reset email: {e}")
        return JSONResponse({"status": "error", "message": "Failed to send email."}, status_code=500)
    
    return JSONResponse({"status": "ok"})

@app.post("/api/reset-password")
async def reset_password(token: str = Form(...), new_password: str = Form(...)):
    """Reset password using valid token."""
    if len(new_password) < 6:
        return JSONResponse({"status": "error", "message": "Password must be at least 6 characters."}, status_code=400)
    
    user = get_user_by_reset_token(token)
    if not user:
        return JSONResponse({"status": "error", "message": "Invalid or expired token."}, status_code=400)
    
    update_user_password(user["email"], new_password)
    clear_reset_token(user["email"])
    
    return JSONResponse({"status": "ok"})

# =========================================================
# LEGACY EMAIL VERIFICATION ROUTES (if you still need them)
# =========================================================
@app.get("/verify-email/{token}")
async def verify_email_route(request: Request, token: str):
    try:
        email = decode_token(token)
    except SignatureExpired:
        return templates.TemplateResponse("pages/signup.html", {
            "request": request,
            "flashes": [{"category": "error", "message": "Verification link expired. Please sign up again."}]
        })
    except BadSignature:
        return templates.TemplateResponse("pages/signup.html", {
            "request": request,
            "flashes": [{"category": "error", "message": "Invalid verification link."}]
        })

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("UPDATE users SET is_verified = 1 WHERE email = %s", (email,))
    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse("/login?verified=1", status_code=302)

@app.post("/resend-verification")
async def resend_verification_route(email: str = Form(...)):
    user = get_user_by_email(email)
    if not user or user["is_verified"] == 1:
        return JSONResponse({"status": "ok"})   # silently ignore

    token = generate_token(email)
    try:
        send_verification_email(email, token)
    except Exception:
        return JSONResponse({"status": "error"}, status_code=500)

    return JSONResponse({"status": "ok"})

# =========================================================
# MAPPING APIs (PERSISTENT – CORRECT VERSION)
# =========================================================
@app.get("/api/companies")
async def get_companies(
    request: Request,
    user: str = Depends(require_login)
):
    user_id = request.session.get("user_id") 
    return {"companies": load_companies(user_id)}

@app.post("/api/companies")
async def create_company(
    request: Request,
    name: str = Form(...),
    user: str = Depends(require_login)
):
    try:
        user_id = request.session.get("user_id")
        add_company(name, user_id)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return {"status": "success"}

@app.delete("/api/companies/{name}")
async def remove_company(
    request: Request,
    name: str,
    user: str = Depends(require_login)
):
    try:
        user_id = request.session.get("user_id")
        delete_company(name, user_id)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return {"status": "deleted"}

@app.get("/api/mapping/{company}")
async def get_company_mapping_api(
    request: Request,
    company: str,
    user: str = Depends(require_login)
):
    try:
        user_id = request.session.get("user_id")
        return get_company_mapping_data(company, user_id)
    except ValueError:
        raise HTTPException(404)

@app.post("/api/mapping/{company}")
async def update_company_mapping(
    request: Request,
    company: str,
    mapping: dict,
    user: str = Depends(require_login)
):
    try:
        user_id = request.session.get("user_id")
        save_company_mapping(company, mapping, user_id)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return {"status": "saved"}

# ================================
# 🔌 CONNECTOR APIs
# ================================
@app.post("/api/add-job/{user_id}")
def add_job(user_id: str, data: dict):
    JOBS.setdefault(user_id, []).append(data)
    return {"ok": True}


@app.get("/api/get-job/{user_id}")
def get_job(user_id: str):

    if JOBS.get(user_id):
        job = JOBS[user_id].pop(0)
        return job

    return {}


@app.post("/api/submit-result/{user_id}")
def submit_result(user_id: str, data: dict):
    RESULTS[user_id] = data
    return {"ok": True}


@app.get("/api/get-result/{user_id}")
def get_result(
    request: Request,
    user_id: str,
    user: str = Depends(require_login),
):
    session_user_id = get_session_user_id(request)
    if session_user_id != str(user_id):
        raise HTTPException(status_code=403, detail="Cannot access another user's result")
    print("CURRENT USER:", user_id)
    return RESULTS.get(user_id, {})

# =========================================================
# PROTECTED APIs (ORDER PRESERVED)
# =========================================================
@app.post("/api/image-to-excel")
async def image_to_excel_api(
    request: Request,
    file: UploadFile,
    company_key: str = Form(...),
    user: str = Depends(require_login)
):
    # --- BUSINESS EVENT LOGGING: OCR STARTED ---
    user_id = request.session.get("user_id")
    username = request.session.get("username", "anonymous")
    start_time = time.perf_counter()
    
    log_admin_event(
        user_id=user_id,
        username=username,
        event_type="ocr_started",
        status_str="success"
    )

    try:
        excel_bytes, output_filename = image_to_excel(
            await file.read(),
            file.filename,
            company_key
        )
        
        # --- BUSINESS EVENT LOGGING: OCR COMPLETED SUCCESS ---
        duration_ms = int((time.perf_counter() - start_time) * 1000)
        
        # Approximate metrics parsing based on image input boundaries
        log_ocr_event(
            user_id=user_id,
            username=username,
            status="success",
            duration_ms=duration_ms,
            file_type=file.filename.split('.')[-1].lower() if '.' in file.filename else 'unknown',
            pages=1, # Base aggregate fallback metric values safely
            rows_generated=0 # Safely derived defaults
        )
        
        return Response(
            content=excel_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={output_filename}"}
        )
    except Exception as ocr_err:
        # --- BUSINESS EVENT LOGGING: OCR EXCEPTION ---
        log_business_error(
            user_id=user_id,
            username=username,
            event_type="ocr",
            error_type="processing_error",
            error_message=str(ocr_err)
        )
        raise ocr_err

@app.post("/api/sheets")
async def get_sheet_names(
    request: Request,
    file: UploadFile,
    user: str = Depends(require_login)
):
    contents = await file.read()
    if file.filename.endswith(".xlsx"):
        wb = openpyxl.load_workbook(BytesIO(contents), read_only=True)
        return {"sheets": wb.sheetnames}
    df = pd.read_excel(BytesIO(contents), sheet_name=None)
    return {"sheets": list(df.keys())}

@app.get("/download-template")
async def download_template(
    request: Request,
    user: str = Depends(require_login)
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"

    headers = [
        'Sr','GSTIN','Recipient Name','Invoice Number',
        'Invoice date','Invoice Value','Taxable Value',
        'IGST','CGST','SGST','Cess'
    ]

    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)

    excel_bytes = io.BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)

    return Response(
        content=excel_bytes.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=invoice_template.xlsx"}
    )

@app.get("/reset-password")
async def reset_password_page(request: Request, token: str = None):
    """Render the login page with the reset token for frontend handling."""
    # Pass the token to the template so JavaScript can pick it up
    return templates.TemplateResponse(
        "pages/login.html",
        {"request": request, "reset_token": token}
    )

from fastapi import Query

@app.get("/api/tally/ledgers")
def api_tally_ledgers(
    request: Request,
    group: str = Query(None),
    user: str = Depends(require_login),
):
    print("📦 API received group:", group)  

    user_id = get_session_user_id(request)
    print("CURRENT USER:", user_id)

    # 1. create XML with group
    xml = build_ledger_xml(group)

    print("📤 XML SENT:\n", xml)  

    # 2. send job (avoid stale results)
    RESULTS.pop(user_id, None)
    JOBS.setdefault(user_id, []).append({"xml": xml})

    # 3. wait for connector response (same pattern as /api/match-party)
    result = None
    import time

    for _ in range(20):  # ~10 seconds max
        result = RESULTS.get(user_id)
        if result:
            break
        time.sleep(0.5)

    if not result:
        return {"status": "waiting"}

    # 4. parse result
    from core.tally_service import parse_ledgers_with_parent

    group_norm = (group or "").strip()

    raw_xml = result.get("data", "") or ""
    try:
        print(
            f"📥 Tally raw XML stats: len={len(raw_xml)}, "
            f"LEDGER={raw_xml.count('<LEDGER')}, "
            f"LEDGERNAME={raw_xml.count('<LEDGERNAME')}, "
            f"NAME={raw_xml.count('<NAME')}, "
            f"LEDGERENTRIES={raw_xml.count('LEDGERENTRIES')}, "
            f"PARENT={raw_xml.count('<PARENT')}, "
            f"PARENTNAME={raw_xml.count('<PARENTNAME')}"
        )
    except Exception:
        pass

    if group_norm and group_norm.lower() != "all":
        parsed = parse_ledgers_with_parent(result.get("data", ""))
        group_norm_lower = group_norm.lower()
        allowed = {
            i["name"]
            for i in parsed
            if i.get("parent", "").strip().lower() == group_norm_lower
        }
        try:
            parents = [i.get("parent", "") for i in parsed]
            parents_clean = [p.strip() for p in parents if p and p.strip()]
            uniq_parents = sorted(
                list({p.lower() for p in parents_clean})
            )[:10]
            eq_count = sum(1 for p in parents_clean if p.lower() == group_norm.lower())
            print(
                f"🧩 parent debug: group={group_norm!r}, parsed={len(parsed)}, "
                f"eq_count={eq_count}, uniq_parents_sample={uniq_parents}"
            )
        except Exception:
            pass
        try:
            parents = [i.get("parent", "") for i in parsed]
            uniq_parents = len(set(p.strip().lower() for p in parents if p.strip()))
            sample_parent = next((p for p in parents if p and p.strip()), "")
            print(
                f"📌 Parent parse stats: parsed={len(parsed)}, uniq_parents={uniq_parents}, sample_parent={sample_parent!r}"
            )
        except Exception:
            pass
        if not allowed:
            allowed = {
                i["name"]
                for i in parsed
                if group_norm_lower in i.get("parent", "").strip().lower()
            }
            print(
                f"🔁 contains-parent fallback: group={group_norm!r}, allowed={len(allowed)}"
            )

        if allowed:
            ledgers = [i["name"] for i in parsed if i["name"] in allowed]
            print(
                f"📦 Group filter fallback applied: group={group_norm}, total={len(parsed)}, allowed={len(allowed)}"
            )
        else:
            ledgers = parse_ledgers(result.get("data", ""))
            print(
                f"⚠️ Group filter fallback: group={group_norm}, total_parsed={len(parsed)}, unfiltered_names={len(ledgers)}"
            )
    else:
        ledgers = parse_ledgers(result.get("data", ""))

    ledgers = list(dict.fromkeys(ledgers))
    return {"status": "ok", "ledgers": ledgers}

@app.post("/api/match-party")
async def match_party(
    request: Request,
    file: UploadFile = File(...),
    tally_group: str = Form(None), 
    sheet_name: str = Form(None),
    manual_columns: str = Form("{}"),
    user: str = Depends(require_login),
):
    # --- BUSINESS EVENT LOGGING: INITIALIZED ---
    start_time = time.perf_counter()
    user_id = request.session.get("user_id")
    username = request.session.get("username", "anonymous")

    try:
        import io
        import json
        import asyncio

        contents = await file.read()
        print("📄 Sheet received from frontend:", sheet_name)

        if sheet_name and sheet_name.strip():
            df = pd.read_excel(io.BytesIO(contents), sheet_name=sheet_name)
        else:
            print("⚠️ No sheet provided, using default Sheet1")
            df = pd.read_excel(io.BytesIO(contents))

        df = df.fillna("")

        # --- BUSINESS EVENT LOGGING: EXCEL FILE LOADED ---
        log_admin_event(
            user_id=user_id,
            username=username,
            event_type="match_file_loaded",
            status_str="success",
            details={
                "rows": len(df),
                "columns": len(df.columns),
                "sheet_name": sheet_name or "Sheet1"
            }
        )

        from core.match_service import (
            detect_party_column,
            detect_gstin_column,
            match_party_names,
            apply_match_results_to_dataframe,
        )
        from core.tally_service import (
            parse_ledgers_with_gstin,
            build_ledger_xml,
        )

        party_col = detect_party_column(df)
        gstin_col = detect_gstin_column(df)

        print("🧠 Detected party column:", party_col)
        print("🧠 Detected GSTIN column:", gantin_col if 'gantin_col' in locals() else gstin_col)

        try:
            manual_cols = json.loads(manual_columns) if isinstance(manual_columns, str) else manual_columns
        except:
            manual_cols = {}

        if manual_cols.get("party"):
            party_col = manual_cols["party"]
            print("🔧 Using manual party column:", party_col)

        if manual_cols.get("gstin"):
            gstin_col = manual_cols["gstin"]
            print("🔧 Using manual GSTIN column:", gstin_col)

        # --- BUSINESS EVENT LOGGING: COLUMNS DETECTED ---
        log_admin_event(
            user_id=user_id,
            username=username,
            event_type="match_columns_detected",
            status_str="success",
            details={
                "party_column": party_col,
                "gstin_column": gstin_col
            }
        )

        if not party_col or not gstin_col:
            print("⚠️ Missing columns - returning manual_required")
            
            # Record structural validation failure
            log_business_error(
                user_id=user_id,
                username=username,
                event_type="match_party",
                error_type="validation_error",
                error_message="GSTIN or Party column mapping could not be explicitly auto-resolved."
            )
            return {
                "status": "manual_required",
                "columns": list(df.columns),
                "party_column": party_col,
                "gstin_column": gstin_col,
            }

        print("🔄 Fetching Tally ledgers...")
        user_id_str = get_session_user_id(request)
        print("CURRENT USER:", user_id_str)

        print("📦 MATCH using group:", tally_group)

        xml = build_ledger_xml(tally_group)
        RESULTS.pop(user_id_str, None) 
        JOBS.setdefault(user_id_str, []).append({"xml": xml})
        print("🧾 JOB ADDED:", JOBS)

        result = None
        for _ in range(20):  
            await asyncio.sleep(0.5)
            result = RESULTS.get(user_id_str)
            if result:
                break

        if not result:
            print("⏳ Waiting for connector response...")
            log_business_error(
                user_id=user_id,
                username=username,
                event_type="match_party",
                error_type="connector_timeout",
                error_message="Tally bridge connector interface timed out waiting for matching parameters ledger streaming."
            )
            return {"status": "waiting"}

        raw_xml = result.get("data", "") or ""
        ledgers, g_map = parse_ledgers_with_gstin(raw_xml)

        group_norm = (tally_group or "").strip()
        if group_norm and group_norm.lower() != "all":
            from core.tally_service import parse_ledgers_with_parent

            parsed = parse_ledgers_with_parent(raw_xml)
            group_norm_lower = group_norm.lower()
            allowed = {
                i["name"]
                for i in parsed
                if i.get("parent", "").strip().lower() == group_norm_lower
            }

            if allowed:
                ledgers = [l for l in ledgers if l in allowed]
                g_map = {
                    gstin: name
                    for gstin, name in g_map.items()
                    if name in allowed
                }
            else:
                allowed = {
                    i["name"]
                    for i in parsed
                    if group_norm_lower in i.get("parent", "").strip().lower()
                }
                if allowed:
                    ledgers = [l for l in ledgers if l in allowed]
                    g_map = {
                        gstin: name
                        for gstin, name in g_map.items()
                        if name in allowed
                    }

        # --- BUSINESS EVENT LOGGING: TALLY LEDGERS STREAMED ---
        log_admin_event(
            user_id=user_id,
            username=username,
            event_type="tally_ledgers_loaded",
            status_str="success",
            details={
                "ledgers_fetched": len(ledgers),
                "tally_group": tally_group or "All"
            }
        )

        print(f"✅ Ledgers fetched: {len(ledgers)}")

        results = match_party_names(
            df=df,
            tally_ledgers=ledgers,
            tally_gstin_map=g_map,
            party_col=party_col,
            gstin_col=gstin_col,
        )

        reviewed_df = apply_match_results_to_dataframe(
            df=df,
            match_results=results,
            party_col=party_col,
        )

        session_id = str(uuid.uuid4())
        MATCH_SESSIONS[session_id] = {
            "reviewed_df": reviewed_df,
            "match_results": results,
            "party_col": party_col,
            "gstin_col": gstin_col,
            "ledger_list": ledgers,
            "sheet_name": sheet_name,
            "source_filename": file.filename,
            "columns": list(df.columns),
        }

        unmatched_count = len(
            [r for r in results if r.get("status") != "matched"]
        )
        matched_count = len(results) - unmatched_count
        
        # Approximate matching classification metadata metrics calculation safely
        exact_count = len([r for r in results if r.get("confidence", 0) == 100 or r.get("status") == "matched"])
        fuzzy_count = matched_count - exact_count

        # --- BUSINESS EVENT LOGGING: MATCH COMPLETE SUCCESS ---
        duration_ms = int((time.perf_counter() - start_time) * 1000)
        log_match_event(
            user_id=user_id,
            username=username,
            status="success",
            duration_ms=duration_ms,
            rows_processed=len(df),
            matched=matched_count,
            unmatched=unmatched_count,
            ledgers_fetched=len(ledgers)
        )
        
        # Log extended internal structured payload safely for cross-dashboard evaluation
        log_admin_event(
            user_id=user_id,
            username=username,
            event_type="match_party_complete",
            status_str="success",
            details={
                "rows_processed": len(df),
                "matched_rows": matched_count,
                "unmatched_rows": unmatched_count,
                "fuzzy_matches": fuzzy_count if fuzzy_count >= 0 else 0,
                "exact_matches": exact_count,
                "duration_ms": duration_ms
            }
        )

        return {
            "status": "ok",
            "session_id": session_id,
            "party_column": party_col,
            "gstin_column": gstin_col,
            "rows": results,
            "ledger_list": ledgers,
            "unmatched_count": unmatched_count,
            "block_convert": unmatched_count > 0,
        }

    except Exception as match_exception:
        print("❌ MATCH ERROR:", str(match_exception))
        
        # --- BUSINESS EVENT LOGGING: UNHANDLED EXCEPTION CORNER ---
        log_business_error(
            user_id=user_id,
            username=username,
            event_type="match_party",
            error_type="system_exception",
            error_message=str(match_exception)
        )
        return {"status": "error", "message": str(match_exception)}


@app.post("/api/apply-corrections")
async def api_apply_corrections(payload: dict):
    try:
        session_id = payload.get("session_id")
        corrections = payload.get("corrections", {})

        if not session_id:
            raise HTTPException(status_code=400, detail="session_id is required")

        session = MATCH_SESSIONS.get(session_id)
        if not session:
            raise HTTPException(status_code=404, detail="Matching session not found")

        reviewed_df = session["reviewed_df"]
        party_col = session["party_col"]

        normalized_corrections = {}
        for k, v in corrections.items():
            try:
                normalized_corrections[int(k)] = v
            except Exception:
                continue

        final_df = apply_corrections_and_build_final_df(
            reviewed_df=reviewed_df,
            corrections=normalized_corrections,
            party_col=party_col,
        )

        session["reviewed_df"] = final_df

        unmatched = [
            row for row in session["match_results"]
            if row.get("status") != "matched" and not normalized_corrections.get(row["row_index"])
        ]

        return {
            "status": "ok",
            "unmatched_count": len(unmatched),
            "can_convert": len(unmatched) == 0,
            "preview_rows": final_df.to_dict(orient="records"),
        }

    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))

@app.post("/api/download-reviewed-excel")
async def api_download_reviewed_excel(payload: dict):
    try:
        session_id = payload.get("session_id")
        corrections = payload.get("corrections", {})

        if not session_id:
            raise HTTPException(status_code=400, detail="session_id is required")

        session = MATCH_SESSIONS.get(session_id)
        if not session:
            raise HTTPException(status_code=404, detail="Matching session not found")

        reviewed_df = session.get("reviewed_df")
        if reviewed_df is None:
            reviewed_df = session.get("df")

        if reviewed_df is None:
            raise HTTPException(status_code=500, detail="No dataframe found in session")

        party_col = session.get("party_col")
        if not party_col:
            raise HTTPException(status_code=500, detail="Party column missing in session")

        normalized_corrections = {}
        for k, v in corrections.items():
            try:
                normalized_corrections[int(k)] = v
            except:
                continue

        print("Corrections:", normalized_corrections)

        final_df = apply_corrections_and_build_final_df(
            reviewed_df=reviewed_df,
            corrections=normalized_corrections,
            party_col=party_col,
        )

        print("Final DF shape:", final_df.shape)
        excel_bytes = export_dataframe_to_excel_bytes(final_df)

        output = BytesIO(excel_bytes)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=reviewed_matching_result.xlsx"
            },
        )

    except HTTPException:
        raise

    except Exception as exc:
        print("❌ ERROR:", str(exc))
        raise HTTPException(status_code=500, detail=str(exc))

@app.post("/api/convert")
async def convert_excel_api(
    request: Request,
    file: UploadFile,
    sheet_name: str = Form(...),
    vtype: str = Form("sale"),
    company: str = Form("Default"),
    column_mapping: str = Form("{}"),
    tally_corrections: str = Form("{}"),
    user: str = Depends(require_login),
):
    """Start Excel→XML conversion as background job.
    
    INSTANT RESPONSE: Returns job_id immediately
    CLIENT POLLS: /api/job-status/{job_id} for progress
    REAL-TIME: WebSocket /ws/job-progress/{job_id} for live updates
    """
    # Generate unique job ID
    job_id = str(uuid.uuid4())
    user_id = request.session.get("user_id")
    username = request.session.get("username", "anonymous")
    start_time = time.perf_counter()
    
    # Initialize job status
    JOB_STATUS[job_id] = {
        "status": "PENDING",
        "progress": 0,
        "message": "Queued for processing...",
        "created_at": datetime.now().isoformat()
    }
    
    if not file.filename.endswith((".xlsx", ".xls")):
        JOB_STATUS[job_id]["status"] = "FAILED"
        JOB_STATUS[job_id]["message"] = "Only Excel files allowed"
        return {"job_id": job_id, "status": "FAILED", "message": "Only Excel files allowed"}

    try:
        mapping_data = json.loads(column_mapping)
    except Exception:
        mapping_data = {}

    try:
        tally_corrections_data = json.loads(tally_corrections)
    except Exception:
        tally_corrections_data = {}

    # Read file into memory
    excel_raw_stream = await file.read()
    
    # Start background processing
    async def process_conversion():
        try:
            JOB_STATUS[job_id]["status"] = "PROCESSING"
            JOB_STATUS[job_id]["progress"] = 5
            JOB_STATUS[job_id]["message"] = "Starting conversion..."
            
            # Log conversion started
            log_admin_event(
                user_id=user_id,
                username=username,
                event_type="convert_started",
                status_str="success"
            )
            
            # Run blocking operation in thread pool (non-blocking)
            def blocking_conversion():
                return excel_to_xml(
                    excel_raw_stream,
                    sheet_name,
                    vtype,
                    company,
                    user_id,
                    column_mapping=mapping_data,
                    tally_corrections=tally_corrections_data
                )
            
            loop = asyncio.get_event_loop()
            # Mark progress: file loaded and queued for processing
            JOB_STATUS[job_id]["progress"] = 20
            JOB_STATUS[job_id]["message"] = "Parsing Excel and generating XML..."

            xml_content, count = await loop.run_in_executor(thread_pool, blocking_conversion)

            # Conversion completed in worker
            JOB_STATUS[job_id]["progress"] = 80
            JOB_STATUS[job_id]["message"] = "Finalizing output..."
            
            # Store result
            RESULTS[job_id] = {
                "content": xml_content,
                "filename": f"{file.filename}_output.xml",
                "count": count
            }

            # Update progress for storing and completion
            JOB_STATUS[job_id]["progress"] = 95
            JOB_STATUS[job_id]["message"] = "Preparing download..."
            
            # Log success
            duration_ms = int((time.perf_counter() - start_time) * 1000)
            log_admin_event(
                user_id=user_id,
                username=username,
                event_type="xml_generated",
                status_str="success",
                details={"rows": count, "voucher_type": vtype}
            )
            
            log_conversion_event(
                user_id=user_id,
                username=username,
                status="success",
                duration_ms=duration_ms,
                rows_processed=count,
                voucher_type=vtype,
                exceptions=0
            )
            
            # Update job status
            JOB_STATUS[job_id]["status"] = "COMPLETED"
            JOB_STATUS[job_id]["progress"] = 100
            JOB_STATUS[job_id]["message"] = f"✅ Conversion complete! {count} rows processed"
            JOB_STATUS[job_id]["completed_at"] = datetime.now().isoformat()
            
        except Exception as e:
            logging.error(f"Conversion error: {str(e)}")
            log_business_error(
                user_id=user_id,
                username=username,
                event_type="convert_xml",
                error_type="conversion_failed",
                error_message=str(e)
            )
            
            JOB_STATUS[job_id]["status"] = "FAILED"
            JOB_STATUS[job_id]["message"] = f"❌ Error: {str(e)}"
            JOB_STATUS[job_id]["error"] = str(e)
    
    # Start background task (fire and forget)
    asyncio.create_task(process_conversion())
    
    # Return immediately with job_id
    return {
        "job_id": job_id,
        "status": "PENDING",
        "message": "Processing started. Check status with /api/job-status/{job_id}",
        "websocket_url": f"/ws/job-progress/{job_id}"
    }

@app.get("/api/job-status/{job_id}")
async def get_job_status(job_id: str):
    """Poll for job status and progress.
    
    Returns: {"status": "PENDING|PROCESSING|COMPLETED|FAILED", "progress": 0-100, "message": "..."}
    """
    if job_id not in JOB_STATUS:
        return {"status": "NOT_FOUND", "message": "Job not found"}
    
    return JOB_STATUS[job_id]


@app.get("/api/job-result/{job_id}")
async def get_job_result(job_id: str):
    """Download converted XML file once job completes.
    
    Returns: XML file or error message
    """
    # Check job status
    if job_id not in JOB_STATUS:
        raise HTTPException(404, "Job not found")
    
    if JOB_STATUS[job_id]["status"] != "COMPLETED":
        raise HTTPException(
            status_code=202,  # Accepted but not ready
            detail={
                "status": JOB_STATUS[job_id]["status"],
                "progress": JOB_STATUS[job_id].get("progress", 0),
                "message": JOB_STATUS[job_id]["message"]
            }
        )
    
    # Get result
    if job_id not in RESULTS:
        raise HTTPException(404, "Result not found")
    
    result = RESULTS[job_id]
    
    return Response(
        content=result["content"],
        media_type="application/xml",
        headers={
            "Content-Disposition": f"attachment; filename={result['filename']}",
            "X-Records-Processed": str(result["count"]),
        },
    )


@app.websocket("/ws/job-progress/{job_id}")
async def websocket_job_progress(websocket: WebSocket, job_id: str):
    """Real-time progress updates via WebSocket.
    
    Sends: {"status": "...", "progress": 0-100, "message": "..."}
    """
    await websocket.accept()
    
    if job_id not in JOB_STATUS:
        await websocket.send_json({"error": "Job not found"})
        await websocket.close()
        return
    
    try:
        # Send initial status
        await websocket.send_json(JOB_STATUS[job_id])
        
        # Keep connection alive and send updates
        last_status = JOB_STATUS[job_id]["status"]
        while True:
            await asyncio.sleep(0.5)  # Check every 500ms
            
            if job_id not in JOB_STATUS:
                break
            
            current_status = JOB_STATUS[job_id]
            
            # Send update only if status changed
            if current_status["status"] != last_status or current_status.get("progress", 0) > 0:
                await websocket.send_json(current_status)
                last_status = current_status["status"]
            
            # Close when job completes
            if current_status["status"] in ["COMPLETED", "FAILED"]:
                await websocket.send_json(current_status)
                break
                
    except WebSocketDisconnect:
        pass
    except Exception as e:
        logging.error(f"WebSocket error: {e}")


# Background cleanup for finished/old jobs to prevent unbounded memory growth
JOB_TTL_SECONDS = 60 * 60  # 1 hour
CLEANUP_INTERVAL_SECONDS = 60  # run cleanup every minute

async def cleanup_old_jobs():
    while True:
        try:
            now = datetime.now()
            expired = []
            for jid, meta in list(JOB_STATUS.items()):
                # Determine timestamp to use
                ts_text = meta.get("completed_at") or meta.get("created_at")
                if not ts_text:
                    continue
                try:
                    ts = datetime.fromisoformat(ts_text)
                except Exception:
                    continue
                age = (now - ts).total_seconds()
                if age > JOB_TTL_SECONDS:
                    expired.append(jid)

            for jid in expired:
                JOB_STATUS.pop(jid, None)
                RESULTS.pop(jid, None)
                logging.info(f"Cleaned up job {jid} after TTL expiry")

        except Exception as e:
            logging.error(f"Error in cleanup_old_jobs: {e}")

        await asyncio.sleep(CLEANUP_INTERVAL_SECONDS)


@app.on_event("startup")
async def start_background_tasks():
    # Ensure created_at exists for any job queued before restart
    for jid, meta in JOB_STATUS.items():
        if "created_at" not in meta:
            meta["created_at"] = datetime.now().isoformat()

    # Launch cleanup loop
    asyncio.create_task(cleanup_old_jobs())


# =========================================================
# MOUNT MULTI-PAGE ENTERPRISE ADMIN SYSTEM ROUTERS
# =========================================================
app.include_router(enterprise_admin_system_router)
app.include_router(business_router)  # <--- NEW BUSINESS ANALYTICS MOUNTED HERE
print("⚡ Enterprise Administration Monitoring Engine fully initialized.")

@app.get("/debug/persistence")
def debug_persistence():
    import os
    
    result = {
        "app_status": "running",
        "database_url_set": bool(os.environ.get("DATABASE_URL")),
        "environment": os.environ.get("RENDER", "not set"),
    }
    
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        
        cur.execute("SELECT EXISTS (SELECT FROM information_schema.tables WHERE table_name = 'users')")
        users_table_exists = cur.fetchone()[0]
        result["users_table_exists"] = users_table_exists
        
        if users_table_exists:
            cur.execute("SELECT COUNT(*) FROM users")
            user_count = cur.fetchone()[0]
            result["user_count"] = user_count
            
            cur.execute("SELECT username FROM users LIMIT 5")
            users = [row[0] for row in cur.fetchall()]
            result["sample_users"] = users
        
        cur.execute("SELECT EXISTS (SELECT FROM information_schema.tables WHERE table_name = 'pending_users')")
        pending_exists = cur.fetchone()[0]
        result["pending_users_table_exists"] = pending_exists
        
        if pending_exists:
            cur.execute("SELECT COUNT(*) FROM pending_users")
            pending_count = cur.fetchone()[0]
            result["pending_count"] = pending_count
        
        cur.close()
        conn.close()
        result["database_connected"] = True
        
    except Exception as e:
        result["database_connected"] = False
        result["database_error"] = str(e)
    
    result["old_sqlite_exists"] = os.path.exists("/data/users.db")
    return result

@app.get("/debug/smtp-test")
async def debug_smtp():
    import socket
    import smtplib
    results = {}

    server = os.environ.get("MAIL_SERVER", "smtp.gmail.com")
    port = int(os.environ.get("MAIL_PORT", "587"))
    username = os.environ.get("MAIL_USERNAME", "")
    password = os.environ.get("MAIL_PASSWORD", "")

    try:
        ip = socket.gethostbyname(server)
        results["dns_resolution"] = {"host": server, "ip": ip, "status": "ok"}
    except Exception as e:
        results["dns_resolution"] = {"error": str(e), "status": "fail"}

    try:
        with smtplib.SMTP(server, port, timeout=10) as smtp:
            smtp.ehlo()
            if port == 587:
                smtp.starttls()
            smtp.ehlo()
            results["smtp_connection"] = {"status": "ok", "banner": str(smtp.ehlo())}
    except Exception as e:
        results["smtp_connection"] = {"error": str(e), "status": "fail"}

    if results.get("smtp_connection", {}).get("status") == "ok" and username and password:
        try:
            with smtplib.SMTP(server, port, timeout=10) as smtp:
                smtp.ehlo()
                if port == 587:
                    smtp.starttls()
                smtp.ehlo()
                smtp.login(username, password)
                results["smtp_login"] = {"status": "ok"}
        except Exception as e:
            results["smtp_login"] = {"error": str(e), "status": "fail"}

    return results